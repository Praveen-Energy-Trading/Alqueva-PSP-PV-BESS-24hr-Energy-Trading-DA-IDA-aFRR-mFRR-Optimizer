"""
backtest_report_exporter.py — write backtest results to Excel.

One sheet of per-day rows plus an aggregate summary block. Output:
<repo_root>/runtime/reports/backtest_<start>_<n>d.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Font

from phase_6_backtesting_and_validation.backtest_engine.backtest_runner import (
    BacktestResult, StochasticComparisonResult,
)


def _repo_root() -> str:
    # this file: <repo>/phase_6_backtesting_and_validation/backtest_excel_reports/ -> up 2
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.abspath(os.path.join(here, os.pardir, os.pardir))


def export_backtest(start_date: str, result: BacktestResult) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest"
    headers = ["date", "feasible", "checker_pass", "objective_eur",
               "realised_revenue_eur", "realised_price_source",
               "realised_afrr_capacity_eur", "realised_mfrr_capacity_eur",
               "realised_reserve_price_source",
               "ida1_feasible", "realised_ida1_revenue_eur", "realised_ida1_price_source",
               "ida2_feasible", "realised_ida2_revenue_eur", "realised_ida2_price_source",
               "ida3_feasible", "realised_ida3_revenue_eur", "realised_ida3_price_source",
               "xbid_feasible", "realised_xbid_revenue_eur", "realised_xbid_price_source",
               "solve_sec", "price_mae", "price_rmse", "price_actual_source",
               "pv_mae", "pv_actual_source", "note"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in result.rows:
        ws.append([r.get(h) for h in headers])

    ws2 = wb.create_sheet("Summary")
    bold = Font(bold=True)
    ws2["A1"] = f"Backtest summary — {start_date}, {result.n_days} days"
    ws2["A1"].font = Font(bold=True, size=13)
    for i, (label, val) in enumerate([
        ("Days", result.n_days),
        ("Feasible", result.n_feasible),
        ("Checker passed", result.n_checker_pass),
        ("Avg objective, forecast-based (EUR)", round(result.avg_objective_eur, 2)),
        ("Real-price days", f"{result.n_real_price_days}/{result.n_days}"),
        ("Avg realised revenue, real OMIE price (EUR)",
         round(result.avg_realised_revenue_eur, 2)
         if result.avg_realised_revenue_eur is not None else "unavailable"),
        ("Real-aFRR-capacity days", f"{result.n_real_afrr_days}/{result.n_days}"),
        ("Avg realised aFRR capacity revenue, real REN price (EUR)",
         round(result.avg_realised_afrr_capacity_eur, 2)
         if result.avg_realised_afrr_capacity_eur is not None else "unavailable"),
        ("Real-mFRR-capacity days", f"{result.n_real_mfrr_days}/{result.n_days}"),
        ("Avg realised mFRR capacity revenue, real REN price (EUR)",
         round(result.avg_realised_mfrr_capacity_eur, 2)
         if result.avg_realised_mfrr_capacity_eur is not None else "unavailable"),
        ("Real-IDA1-price days", f"{result.n_real_ida1_price_days}/{result.n_days}"),
        ("Avg realised IDA1 revenue, real OMIE price (EUR)",
         round(result.avg_realised_ida1_revenue_eur, 2)
         if result.avg_realised_ida1_revenue_eur is not None else "unavailable"),
        ("Real-IDA2-price days", f"{result.n_real_ida2_price_days}/{result.n_days}"),
        ("Avg realised IDA2 revenue, real OMIE price (EUR)",
         round(result.avg_realised_ida2_revenue_eur, 2)
         if result.avg_realised_ida2_revenue_eur is not None else "unavailable"),
        ("Real-IDA3-price days", f"{result.n_real_ida3_price_days}/{result.n_days}"),
        ("Avg realised IDA3 revenue, real OMIE price (EUR)",
         round(result.avg_realised_ida3_revenue_eur, 2)
         if result.avg_realised_ida3_revenue_eur is not None else "unavailable"),
        ("Real-XBID-price days (window W1)", f"{result.n_real_xbid_price_days}/{result.n_days}"),
        ("Avg realised XBID revenue, real OMIE price (EUR)",
         round(result.avg_realised_xbid_revenue_eur, 2)
         if result.avg_realised_xbid_revenue_eur is not None else "unavailable"),
        ("Avg solve (s)", round(result.avg_solve_sec, 3)),
        ("Avg price MAE (EUR/MWh)", round(result.avg_price_mae, 2)),
        ("Avg PV MAE (MW) — PV actual always synthetic", round(result.avg_pv_mae, 4)),
        ("NOTE — out of scope", "Activation/imbalance revenue excluded: activated MW is "
         "always internally-simulated (no real REN/SCADA telemetry loader exists), so "
         "it cannot meet the real-price x real-quantity standard used above. IDA3 has "
         "only 7 real archived dates today (coverage from 2026-08-15) -- genuinely thin, "
         "reported as such rather than padded. XBID is backtested at window W1 only "
         "(production supports 6 continuous-intraday check windows)."),
    ], start=3):
        ws2[f"A{i}"] = label; ws2[f"A{i}"].font = bold
        ws2[f"B{i}"] = val

    # --- Risk sheet ---
    if result.risk is not None:
        wr = wb.create_sheet("Risk")
        wr["A1"] = "Portfolio Risk Metrics"
        wr["A1"].font = Font(bold=True, size=13)
        wr["A2"] = f"Source: {result.risk.n_days} feasible backtest days  |  " \
                   f"Bootstrap n=10,000  |  alpha=95% and 99%"

        risk_rows = [
            ("", ""),
            ("--- P&L Distribution ---", ""),
            ("Mean daily P&L (EUR)",              result.risk.mean_pnl_eur),
            ("Std daily P&L (EUR)",               result.risk.std_pnl_eur),
            ("Min daily P&L (EUR)",               result.risk.min_pnl_eur),
            ("Max daily P&L (EUR)",               result.risk.max_pnl_eur),
            ("", ""),
            ("--- Historical Simulation ---", ""),
            ("VaR(95%)  — 5th-pct P&L (EUR)",    result.risk.var_95_eur),
            ("CVaR(95%) — Expected Shortfall (EUR)", result.risk.cvar_95_eur),
            ("VaR(99%)  — 1st-pct P&L (EUR)",    result.risk.var_99_eur),
            ("CVaR(99%) — Expected Shortfall (EUR)", result.risk.cvar_99_eur),
            ("", ""),
            ("--- Monte Carlo Bootstrap (VaR 95%) ---", ""),
            ("VaR(95%)  mean  (EUR)",             result.risk.var_95_mean),
            ("VaR(95%)  std   (EUR)  ± CI",       result.risk.var_95_std),
            ("CVaR(95%) mean  (EUR)",             result.risk.cvar_95_mean),
            ("CVaR(95%) std   (EUR)  ± CI",       result.risk.cvar_95_std),
            ("", ""),
            ("--- Risk-Adjusted ---", ""),
            ("Sharpe ratio (annualised, rf=0)",   result.risk.sharpe_ratio),
            ("Max drawdown (EUR)",                result.risk.max_drawdown_eur),
        ]
        for i, (label, val) in enumerate(risk_rows, start=4):
            wr[f"A{i}"] = label
            if label.startswith("---"):
                wr[f"A{i}"].font = Font(bold=True)
            if val != "":
                wr[f"B{i}"] = val

    # --- Operational analytics sheet (per-day summary) ---
    feasible_rows = [r for r in result.rows if r.get("feasible") and r.get("ops")]
    if feasible_rows:
        wo = wb.create_sheet("Operational")
        ops_headers = [
            "date", "turbine_hours_total", "pump_hours_total",
            "turbine_starts_total", "pump_starts_total",
            "turb_avg_run_h", "turb_max_run_h",
            "pump_avg_run_h", "pump_max_run_h",
            "turb_hours_top25pct_price", "pump_hours_bot25pct_price",
            "bess_charge_hours", "bess_discharge_hours",
            "avg_units_turbining", "avg_units_pumping",
        ]
        wo.append(ops_headers)
        for c in wo[1]: c.font = Font(bold=True)
        for r in feasible_rows:
            ops = r.get("ops", {})
            wo.append([r["date"]] + [ops.get(k, "") for k in ops_headers[1:]])

        # Temporal split sheet
        wt = wb.create_sheet("Temporal")
        tmp_headers = ["date", "band",
                       "hours", "turbine_pct", "pump_pct",
                       "avg_net_mw", "avg_profit_eur_h", "avg_price_eur_mwh"]
        wt.append(tmp_headers)
        for c in wt[1]: c.font = Font(bold=True)
        for r in feasible_rows:
            tmp = r.get("tmp", {})
            for band in ("night", "morning", "afternoon", "evening"):
                bd = tmp.get(band, {})
                if not bd:
                    continue
                wt.append([r["date"], band] + [bd.get(k, "") for k in tmp_headers[2:]])

        # Extended KPI sheet
        we = wb.create_sheet("KPI_Extended")
        eco_headers = [
            "date",
            "turbine_capacity_factor_pct", "pump_capacity_factor_pct",
            "bess_discharge_cf_pct", "pv_utilisation_pct",
            "avg_turbine_efficiency_pct", "avg_pump_efficiency_pct",
            "reservoir_fill_end_pct",
            "head_min_m", "head_max_m", "head_range_m",
            "da_revenue_share_pct", "frr_revenue_share_pct",
            "energy_revenue_eur",
        ]
        we.append(eco_headers)
        for c in we[1]: c.font = Font(bold=True)
        for r in feasible_rows:
            eco = r.get("eco", {})
            we.append([r["date"]] + [eco.get(k, "") for k in eco_headers[1:]])

    out_dir = os.path.join(_repo_root(), "runtime", "reports")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"backtest_{start_date}_{result.n_days}d.xlsx")
    wb.save(path)
    return path


def export_risk_comparison(start_date: str, result: StochasticComparisonResult) -> str:
    """Write the EV-vs-CVaR realized-outcome comparison to Excel: one sheet
    of per-day realized revenue under each strategy, plus a summary block
    with both strategies' risk statistics (VaR/CVaR/Sharpe/drawdown) side
    by side, the cost-of-risk-aversion, and the tail-improvement figure.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    headers = ["date", "price_source", "feasible_ev", "realised_ev_eur",
               "feasible_cvar", "realised_cvar_eur"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in result.rows:
        ws.append([r.get(h) for h in headers])

    ws2 = wb.create_sheet("Summary")
    bold = Font(bold=True)
    ws2["A1"] = f"EV vs CVaR realized-outcome comparison — {start_date}, {result.n_days} days"
    ws2["A1"].font = Font(bold=True, size=13)
    rows = [
        ("Days", result.n_days),
        ("Real-price days used", f"{result.n_real_days}/{result.n_days}"),
        ("Mean realised revenue, expected-value strategy (EUR)",
         round(result.mean_realised_ev_eur, 2) if result.mean_realised_ev_eur is not None else "unavailable"),
        ("Mean realised revenue, CVaR-averse strategy (EUR)",
         round(result.mean_realised_cvar_eur, 2) if result.mean_realised_cvar_eur is not None else "unavailable"),
        ("Cost of risk aversion (EV mean - CVaR mean, EUR)",
         round(result.cost_of_risk_aversion_eur, 2) if result.cost_of_risk_aversion_eur is not None else "unavailable"),
        ("Tail improvement (CVaR's realized CVaR95 - EV's realized CVaR95, EUR)",
         round(result.tail_improvement_eur, 2) if result.tail_improvement_eur is not None else "unavailable"),
    ]
    for i, (label, val) in enumerate(rows, start=3):
        ws2[f"A{i}"] = label; ws2[f"A{i}"].font = bold
        ws2[f"B{i}"] = val

    if result.risk_ev is not None and result.risk_cvar is not None:
        wr = wb.create_sheet("RiskComparison")
        wr["A1"] = "Realized-outcome risk metrics — EV vs CVaR strategy"
        wr["A1"].font = Font(bold=True, size=13)
        wr["A2"] = "EV"; wr["A2"].font = bold
        wr["B2"] = "CVaR"; wr["B2"].font = bold
        metric_rows = [
            ("Mean daily P&L (EUR)", "mean_pnl_eur"),
            ("Std daily P&L (EUR)", "std_pnl_eur"),
            ("Min daily P&L (EUR)", "min_pnl_eur"),
            ("Max daily P&L (EUR)", "max_pnl_eur"),
            ("VaR(95%) (EUR)", "var_95_eur"),
            ("CVaR(95%) (EUR)", "cvar_95_eur"),
            ("VaR(99%) (EUR)", "var_99_eur"),
            ("CVaR(99%) (EUR)", "cvar_99_eur"),
            ("Sharpe ratio (annualised)", "sharpe_ratio"),
            ("Max drawdown (EUR)", "max_drawdown_eur"),
        ]
        wr["C1"] = "Metric"; wr["C1"].font = bold
        for i, (label, field_name) in enumerate(metric_rows, start=3):
            wr[f"C{i}"] = label
            wr[f"A{i}"] = getattr(result.risk_ev, field_name)
            wr[f"B{i}"] = getattr(result.risk_cvar, field_name)

    out_dir = os.path.join(_repo_root(), "runtime", "reports")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"risk_comparison_{start_date}_{result.n_days}d.xlsx")
    wb.save(path)
    return path
